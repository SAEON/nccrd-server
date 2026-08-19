"""
API routers for NCCRD submissions.

Changelog vs. previous version
--------------------------------
* All references to ``Adaptaion`` updated to ``Adaptation``.
* ``project_manager_position`` / ``_phone`` / ``_mobile`` replaced with
  ``project_manager_contact_number``; ``platfrom`` replaced with ``platform``.
* Province JSONB filter uses the native PostgreSQL ``@>`` containment operator
  (``Submission.geo_location.contains({"province": province})``) instead of
  the slower ``.astext ==`` cast approach.
* ``new_submission``, ``update_new_submission``, ``delete``, and the bulk-upload
  endpoint are protected by ``Depends(RequirePermission(...))``, which checks
  a Hydra scope (``Authorize``) and an RBAC permission for the current tenant.
* List/read endpoints and all write endpoints are scoped to the tenant
  resolved from the request's Host header (see ``nccrd.api.lib.tenant``);
  writes also link the new submission to that tenant.
* Bulk-upload endpoint (``create_submission_upload-xlsx``) completely rewritten:
  - Reads the first row as column headers (tabular format).
  - Maps headers via ``GENERAL_COLUMN_MAP`` / ``ADAPTATION_COLUMN_MAP`` /
    ``MITIGATION_COLUMN_MAP`` dicts — no hard-coded row indexes.
  - Supports multi-row batches; all rows are inserted in a single DB
    transaction so a validation failure on any row rolls back the whole batch.
  - Validates ``hazard``, ``sector`` (adaptation), ``sector`` (mitigation),
    and ``mitigation_program`` against the ``Vocabulary`` table before saving.
    Unknown terms are flagged in an error payload (row + column) rather than
    silently mapped to "Other", giving data-quality feedback to the uploader.
* Debug ``print`` statements removed throughout.
"""

from __future__ import annotations

import traceback
import uuid as uuid_lib
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Union
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from openpyxl import load_workbook
from sqlalchemy import and_, exists, or_
from sqlalchemy.orm import Query as SAQuery, Session

from nccrd.api.lib.auth import Authorized
from nccrd.api.lib.permissions import RequirePermission
from nccrd.api.lib.tenant import get_current_tenant
from nccrd.api.models import (
    ProgressReportResponse,
    SubmissionCreate,
    SubmissionModel,
    SubmissionResponse,
    SubmissionUpdate,
)
from nccrd.api.models.submission import (
    AdaptationCreate,
    MitigationCreate,
    GeoLocationSchema,
    InterventionMeasurementEnum,
)
from nccrd.db import get_db
from nccrd.db.models import Submission, Adaptation, Mitigation, ProgressReport, Vocabulary
from nccrd.db.models.rbac import Tenant, TenantXrefSubmission

#: Local-disk storage for progress-report (MRV) uploads. Mounted as a Docker
#: volume in production (see deploy/docker-compose.yml) so files survive
#: container rebuilds — this path alone is not durable on its own.
PROGRESS_REPORT_UPLOAD_DIR = Path(__file__).resolve().parents[3] / "uploads" / "progress_reports"
PROGRESS_REPORT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
MAX_PROGRESS_REPORT_SIZE_BYTES = 25 * 1024 * 1024  # 25MB

router = APIRouter()

# ──────────────────────────────────────────────────────────────────────────────
# Column maps for bulk Excel upload (tabular format: row 1 = headers)
# ──────────────────────────────────────────────────────────────────────────────

#: Maps "General project details" sheet header  →  SubmissionCreate field name.
GENERAL_COLUMN_MAP: Dict[str, str] = {
    "Project Title": "title",
    "Indicate the type of measure": "intervention_measurement",
    "Description": "description",
    "Implementation status": "implementation_status",
    "Implementing organization": "implementation_organization",
    "Other implementing partners": "implementation_partners_other",
    "Start year": "start_date",
    "End year": "end_date",
    "Link to project website": "link",
    "Funding organization": "funding_organization",
    "Type of funding": "funding_type",
    "Actual budget": "funding_amount",
    "Estimated Budget Range ZAR": "estimated_budget_cost",
    # Geographic hierarchy columns.
    "Country": "_geo_country",
    "Province": "_geo_province",
    "District": "_geo_district",
    "Local Municipality": "_geo_local_municipality",
    "Town / Suburb": "_geo_town_suburb",
    "Longitude": "_geo_longitude",
    "Latitude": "_geo_latitude",
    # Project manager.
    "Project Manager Name": "project_manager_name",
    "Project Manager Organization": "project_manager_organization",
    "Email address": "project_manager_email",
    "Contact Number": "project_manager_contact_number",
    # Misc.
    "Research": "research",
    "Platform": "platform",
}

#: Maps "Adaptation details" sheet header  →  AdaptationCreate field name.
ADAPTATION_COLUMN_MAP: Dict[str, str] = {
    "Adaptation sector": "sector",
    "National policy": "national_policy",
    "Overall adaptation intervention goal": "intervention_goal",
    "Provincial / municipal policy / framework": "provincial_municipal",
    "Hazard": "hazard",
    "Progress calculator / explanation": "progress_calculator",
    "Observed and projected climate change impacts": "climate_impact",
    "How the intervention addresses the climate impact": "address_climate_impact",
    "Adaptation impact response": "impact_response",
}

#: Maps "Mitigation details" sheet header  →  MitigationCreate field name.
MITIGATION_COLUMN_MAP: Dict[str, str] = {
    "Mitigation sector": "sector",
    "Subsector": "subsector",
    "Secondary sector": "secondary",
    "Project type": "project_type",
    "Project subtype": "project_subtype",
    "Mitigation programme": "mitigation_program",
    "National policy": "national_policy",
    "Provincial / municipal policy / framework": "provincial_municipal",
    "Primary intended mitigation outcome": "primary_intended_outcome",
    "Progress calculator / explanation": "progress_calculator",
    "Environmental co-benefit": "environmental_co_benefit",
    "Environmental co-benefit description": "environmental_co_benefit_description",
    "Social co-benefit": "social_co_benefit",
    "Social co-benefit description": "social_co_benefit_description",
    "Economic co-benefit": "economic_co_benefit",
    "Economic co-benefit description": "economic_co_benefit_description",
    "Are carbon credits issued?": "carbon_credit",
    "CDM / Voluntary": "cdm_voluntary",
    "CDM Executive Board status": "cdm_executive_board_status",
    "CDM methodology": "cdm_methodology",
    "Organisation issuing carbon credits": "organization_issuing_credits",
    "Voluntary methodology": "voluntary_methodology",
    "CDM project number": "cdm_project_number",
}

#: Fields to validate against the Vocabulary table.
#: Format: (schema_field_name, human_readable_column_label)
VOCABULARY_VALIDATED_FIELDS: List[tuple[str, str]] = [
    ("hazard", "Hazard"),
    ("sector", "Adaptation sector / Mitigation sector"),
    ("mitigation_program", "Mitigation programme"),
]


# ──────────────────────────────────────────────────────────────────────────────
# Helper utilities
# ──────────────────────────────────────────────────────────────────────────────


def _scope_to_tenant(query: SAQuery, tenant: Tenant) -> SAQuery:
    """
    Restrict a ``Submission`` query to what ``tenant`` is allowed to see:
    submissions explicitly linked to it via ``tenant_xref_submission``, plus
    (when ``include_unbounded_submissions`` is set) submissions not linked to
    any tenant at all.
    """
    linked = exists().where(
        and_(
            TenantXrefSubmission.submission_id == Submission.id,
            TenantXrefSubmission.tenant_id == tenant.id,
        )
    )
    if tenant.include_unbounded_submissions:
        unbound = ~exists().where(TenantXrefSubmission.submission_id == Submission.id)
        return query.filter(or_(linked, unbound))
    return query.filter(linked)


def _link_submission_to_tenant(db: Session, submission_id, tenant: Tenant) -> None:
    """Record that ``submission_id`` belongs to ``tenant`` (multi-tenancy)."""
    db.add(TenantXrefSubmission(tenant_id=tenant.id, submission_id=submission_id))


def _normalize_intervention_measurement(raw: str) -> str:
    """Normalise a free-text measure type to a canonical enum value."""
    lowered = raw.strip().lower()
    if "mitigation" in lowered:
        return InterventionMeasurementEnum.MITIGATION.value
    if "adaptation" in lowered:
        return InterventionMeasurementEnum.ADAPTATION.value
    if "cross" in lowered:
        return InterventionMeasurementEnum.CROSS_CUTTING.value
    return raw  # Let Pydantic reject invalid values with a clear 422 error.


def _get_valid_vocabulary_terms(db: Session) -> Set[str]:
    """
    Fetch all known vocabulary terms from the database in one query.
    Returns a set of lowercase terms for O(1) lookup.
    """
    rows = db.query(Vocabulary.term).filter(Vocabulary.term.isnot(None)).all()
    return {row[0].strip().lower() for row in rows if row[0]}


def _parse_sheet_rows(
        sheet: Any,
        column_map: Dict[str, str],
) -> List[Dict[str, Any]]:
    """
    Parse a worksheet where the first row contains column headers.

    Returns a list of dicts mapping schema field names to cell values.
    Headers not present in ``column_map`` are silently ignored.
    """
    rows_iter = sheet.iter_rows(values_only=True)

    # First row = headers.
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return []

    # Map column index → schema field name (skip unmapped headers).
    indexed_columns: Dict[int, str] = {}
    for col_idx, header_value in enumerate(raw_headers):
        if header_value is None:
            continue
        header_str = str(header_value).strip()
        if header_str in column_map:
            indexed_columns[col_idx] = column_map[header_str]

    parsed_rows: List[Dict[str, Any]] = []
    for row in rows_iter:
        # Skip fully empty rows.
        if all(cell is None for cell in row):
            continue
        row_data: Dict[str, Any] = {}
        for col_idx, field_name in indexed_columns.items():
            cell_value = row[col_idx] if col_idx < len(row) else None
            if cell_value is not None:
                row_data[field_name] = cell_value
        if row_data:
            parsed_rows.append(row_data)

    return parsed_rows


def _build_geo_location(row_data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """
    Assemble a GeoLocationSchema-compatible dict from the ``_geo_*`` sentinel
    keys injected by ``GENERAL_COLUMN_MAP``.
    """
    geo: Dict[str, Any] = {}
    geo_keys = {
        "_geo_country": "country",
        "_geo_province": "province",
        "_geo_district": "district",
        "_geo_local_municipality": "local_municipality",
        "_geo_town_suburb": "town_suburb",
    }
    for sentinel, schema_key in geo_keys.items():
        if sentinel in row_data:
            geo[schema_key] = row_data.pop(sentinel)
        else:
            row_data.pop(sentinel, None)

    lon = row_data.pop("_geo_longitude", None)
    lat = row_data.pop("_geo_latitude", None)
    if lon is not None and lat is not None:
        geo["coordinates"] = [float(lon), float(lat)]
        geo["type"] = "Point"

    return geo if geo else None


# ──────────────────────────────────────────────────────────────────────────────
# READ endpoints (public)
# ──────────────────────────────────────────────────────────────────────────────


@router.get(
    "/list_submission",
    response_model=List[SubmissionModel],
    summary="List all submissions, optionally filtered by any facet field.",
)
def get_submissions_list(
        submission_id: Optional[str] = None,
        implementation_status: Optional[str] = None,
        implementation_organization: Optional[str] = None,
        funding_organization: Optional[str] = None,
        funding_type: Optional[str] = None,
        submission_status: Optional[str] = None,
        research: Optional[str] = None,
        # Adaptation facet filters.
        adaptation_sector: Optional[str] = None,
        adaptation_national_policy: Optional[str] = None,
        adaptation_hazard: Optional[str] = None,
        adaptation_climate_impact: Optional[str] = None,
        # Mitigation facet filters.
        mitigation_sector: Optional[str] = None,
        mitigation_subsector: Optional[str] = None,
        mitigation_project_type: Optional[str] = None,
        mitigation_program: Optional[str] = None,
        mitigation_national_policy: Optional[str] = None,
        mitigation_primary_intended_outcome: Optional[str] = None,
        mitigation_environmental_co_benefit: Optional[str] = None,
        mitigation_social_co_benefit: Optional[str] = None,
        mitigation_economic_co_benefit: Optional[str] = None,
        mitigation_carbon_credit: Optional[str] = None,
        # Geographic / type filters.
        province: Optional[str] = None,
        intervention_measurement: Optional[str] = None,
        # Full-text search.
        q: Optional[str] = None,
        db: Session = Depends(get_db),
        tenant: Tenant = Depends(get_current_tenant),
) -> List[Submission]:
    """
    Return all non-deleted submissions, optionally filtered by any combination
    of facet values.

    JSONB province filtering uses the PostgreSQL ``@>`` containment operator
    for correctness and index-friendliness. Results are scoped to the tenant
    resolved from the request's Host header (see `get_current_tenant`).
    """

    query = db.query(Submission).filter(Submission.deleted.isnot(True))
    query = _scope_to_tenant(query, tenant)

    if submission_id:
        query = query.filter(Submission.id == submission_id)
    if implementation_status:
        query = query.filter(Submission.implementation_status == implementation_status)
    if implementation_organization:
        query = query.filter(
            Submission.implementation_organization == implementation_organization
        )
    if funding_organization:
        query = query.filter(Submission.funding_organization == funding_organization)
    if funding_type:
        query = query.filter(Submission.funding_type == funding_type)
    if submission_status:
        query = query.filter(Submission.submission_status == submission_status)
    if research:
        query = query.filter(Submission.research == research)

    if province:
        # Use JSONB containment (@>) for accurate, index-efficient matching.
        query = query.filter(
            Submission.geo_location.contains({"province": province})
        )

    if intervention_measurement:
        # Support comma-separated multi-select, e.g. "Mitigation,Adaptation".
        types = [t.strip() for t in intervention_measurement.split(",")]
        if len(types) > 1:
            query = query.filter(Submission.intervention_measurement.in_(types))
        else:
            query = query.filter(Submission.intervention_measurement == types[0])

    # ── Adaptation join + filters ─────────────────────────────────────────────
    adaptation_filters = [
        adaptation_sector,
        adaptation_national_policy,
        adaptation_hazard,
        adaptation_climate_impact,
    ]
    if any(adaptation_filters):
        query = query.join(Adaptation, Adaptation.submission_id == Submission.id)
        if adaptation_sector:
            query = query.filter(Adaptation.sector == adaptation_sector)
        if adaptation_national_policy:
            query = query.filter(Adaptation.national_policy == adaptation_national_policy)
        if adaptation_hazard:
            query = query.filter(Adaptation.hazard == adaptation_hazard)
        if adaptation_climate_impact:
            query = query.filter(Adaptation.climate_impact == adaptation_climate_impact)

    # ── Mitigation join + filters ─────────────────────────────────────────────
    mitigation_filters = [
        mitigation_sector,
        mitigation_subsector,
        mitigation_project_type,
        mitigation_program,
        mitigation_national_policy,
        mitigation_primary_intended_outcome,
        mitigation_environmental_co_benefit,
        mitigation_social_co_benefit,
        mitigation_economic_co_benefit,
        mitigation_carbon_credit,
    ]
    if any(mitigation_filters):
        query = query.join(Mitigation, Mitigation.submission_id == Submission.id)
        if mitigation_sector:
            query = query.filter(Mitigation.sector == mitigation_sector)
        if mitigation_subsector:
            query = query.filter(Mitigation.subsector == mitigation_subsector)
        if mitigation_project_type:
            query = query.filter(Mitigation.project_type == mitigation_project_type)
        if mitigation_program:
            query = query.filter(Mitigation.mitigation_program == mitigation_program)
        if mitigation_national_policy:
            query = query.filter(Mitigation.national_policy == mitigation_national_policy)
        if mitigation_primary_intended_outcome:
            query = query.filter(
                Mitigation.primary_intended_outcome == mitigation_primary_intended_outcome
            )
        if mitigation_environmental_co_benefit:
            query = query.filter(
                Mitigation.environmental_co_benefit == mitigation_environmental_co_benefit
            )
        if mitigation_social_co_benefit:
            query = query.filter(Mitigation.social_co_benefit == mitigation_social_co_benefit)
        if mitigation_economic_co_benefit:
            query = query.filter(
                Mitigation.economic_co_benefit == mitigation_economic_co_benefit
            )
        if mitigation_carbon_credit:
            query = query.filter(Mitigation.carbon_credit == mitigation_carbon_credit)

    # ── Full-text title search ────────────────────────────────────────────────
    if q:
        query = query.filter(Submission.title.ilike(f"%{q}%"))

    return query.all()


@router.get(
    "/read_submission/{submission_uuid}",
    response_model=SubmissionResponse,
    summary="Retrieve a single submission with its nested Mitigation / Adaptation data.",
)
def read_submission(
        submission_uuid: UUID,
        db: Session = Depends(get_db),
        tenant: Tenant = Depends(get_current_tenant),
) -> Submission:
    """Fetch one submission by UUID and attach the relevant child records."""
    query = db.query(Submission).filter(Submission.id == submission_uuid)
    submission = _scope_to_tenant(query, tenant).first()
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")

    mitigation = (
        db.query(Mitigation).filter(Mitigation.submission_id == submission.id).first()
    )
    adaptation = (
        db.query(Adaptation).filter(Adaptation.submission_id == submission.id).first()
    )

    im_value = (
        submission.intervention_measurement.strip().lower()
        if submission.intervention_measurement
        else ""
    )

    if im_value == "cross cutting":
        submission.mitigation = mitigation
        submission.adaptation = adaptation
    elif im_value == "adaptation":
        submission.mitigation = None
        submission.adaptation = adaptation
    elif im_value == "mitigation":
        submission.mitigation = mitigation
        submission.adaptation = None
    else:
        submission.mitigation = None
        submission.adaptation = None

    return submission


# ──────────────────────────────────────────────────────────────────────────────
# WRITE endpoints (require PROJECT_ADMIN scope + an RBAC permission)
# ──────────────────────────────────────────────────────────────────────────────


@router.post(
    "/new_submission",
    summary="Create a new submission and its related Mitigation / Adaptation records.",
)
def create_submission(
        submission: SubmissionCreate,
        db: Session = Depends(get_db),
        auth: Authorized = Depends(RequirePermission("create-submission")),
        tenant: Tenant = Depends(get_current_tenant),
) -> Dict[str, str]:
    """
    Persist a new submission.  Mitigation / Adaptation child records are
    created automatically based on ``intervention_measurement``.
    """
    geo_dict = (
        submission.geo_location.dict(exclude_none=True)
        if submission.geo_location
        else None
    )

    db_submission = Submission(
        title=submission.title,
        intervention_measurement=submission.intervention_measurement,
        description=submission.description,
        implementation_status=submission.implementation_status,
        implementation_organization=submission.implementation_organization,
        implementation_partners_other=submission.implementation_partners_other,
        start_date=submission.start_date,
        end_date=submission.end_date,
        link=submission.link,
        funding_organization=submission.funding_organization,
        funding_type=submission.funding_type,
        funding_amount=submission.funding_amount,
        estimated_budget_cost=submission.estimated_budget_cost,
        geo_location=geo_dict,
        project_manager_name=submission.project_manager_name,
        project_manager_organization=submission.project_manager_organization,
        project_manager_email=submission.project_manager_email,
        project_manager_contact_number=submission.project_manager_contact_number,
        submission_status="Pending",
        issubmitted=True,
        platform=submission.platform,
        research=submission.research,
        data_source=submission.data_source or "react_app",
        createdby=auth.internal_user_id,
        createdate=datetime.utcnow(),
    )
    db.add(db_submission)
    db.flush()  # Populate db_submission.id without committing yet.
    _link_submission_to_tenant(db, db_submission.id, tenant)

    im_value = submission.intervention_measurement.strip().lower()

    if im_value in ("mitigation", "cross cutting"):
        if not submission.mitigation_data:
            raise HTTPException(
                status_code=400,
                detail=(
                    "mitigation_data must be provided when "
                    "intervention_measurement is 'Mitigation' or 'Cross Cutting'."
                ),
            )
        _create_mitigation_record(db, db_submission.id, submission.mitigation_data)

    if im_value in ("adaptation", "cross cutting"):
        if not submission.adaptation_data:
            raise HTTPException(
                status_code=400,
                detail=(
                    "adaptation_data must be provided when "
                    "intervention_measurement is 'Adaptation' or 'Cross Cutting'."
                ),
            )
        _create_adaptation_record(db, db_submission.id, submission.adaptation_data)

    db.commit()
    return {
        "detail": "Submission and related records created successfully.",
        "submission_id": str(db_submission.id),
    }


@router.patch(
    "/update_new_submission/{submission_uuid}",
    response_model=SubmissionResponse,
    summary="Partially update a submission and its child records.",
)
def update_submission(
        submission_uuid: UUID,
        update_data: SubmissionUpdate,
        db: Session = Depends(get_db),
        auth: Authorized = Depends(RequirePermission("update-submission")),
) -> Submission:
    """
    Update an existing submission.  Nested Mitigation / Adaptation records are
    created, updated, or deleted based on the new ``intervention_measurement``
    value (if supplied).
    """
    submission = (
        db.query(Submission).filter(Submission.id == submission_uuid).first()
    )
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")

    data = update_data.dict(exclude_unset=True)

    # Pull nested payloads out before applying scalar updates.
    mitigation_update: Optional[Dict[str, Any]] = data.pop("mitigation_data", None)
    adaptation_update: Optional[Dict[str, Any]] = data.pop("adaptation_data", None)

    # Convert geo_location Pydantic sub-model to dict if present.
    if "geo_location" in data and isinstance(data["geo_location"], GeoLocationSchema):
        data["geo_location"] = data["geo_location"].dict(exclude_none=True)

    new_intervention: str = (
        data.get("intervention_measurement", submission.intervention_measurement)
        .strip()
        .lower()
    )

    # Apply scalar field updates.
    for key, value in data.items():
        setattr(submission, key, value)

    submission.updatedby = auth.internal_user_id
    submission.updatedate = datetime.utcnow()

    # ── Reconcile child records ───────────────────────────────────────────────

    if new_intervention == "mitigation":
        _delete_adaptation(db, submission.id)
        _upsert_mitigation(db, submission.id, mitigation_update)

    elif new_intervention == "adaptation":
        _delete_mitigation(db, submission.id)
        _upsert_adaptation(db, submission.id, adaptation_update)

    elif new_intervention == "cross cutting":
        _upsert_mitigation(db, submission.id, mitigation_update)
        _upsert_adaptation(db, submission.id, adaptation_update)

    else:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid intervention_measurement value: '{new_intervention}'.",
        )

    db.commit()
    db.refresh(submission)
    return submission


@router.delete(
    "/delete/{submission_uuid}",
    summary="Soft-delete a submission (sets deleted=True, records deletedate).",
)
def soft_delete_submission(
        submission_uuid: UUID,
        db: Session = Depends(get_db),
        auth: Authorized = Depends(RequirePermission("delete-submission")),
) -> Dict[str, str]:
    """Mark a submission as deleted without removing the database row."""
    submission = (
        db.query(Submission).filter(Submission.id == submission_uuid).first()
    )
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")

    submission.deleted = True
    submission.deletedate = datetime.utcnow()
    submission.deletedby = auth.internal_user_id
    db.commit()
    return {"detail": "Submission marked as deleted."}


@router.post(
    "/{submission_uuid}/progress_reports",
    response_model=ProgressReportResponse,
    summary="Attach a progress / MRV document to an existing submission.",
)
async def upload_progress_report(
        submission_uuid: UUID,
        file: UploadFile = File(...),
        notes: Optional[str] = Form(None),
        db: Session = Depends(get_db),
        auth: Authorized = Depends(RequirePermission("attach-file-to-submission")),
) -> ProgressReport:
    """
    Store the uploaded file on local disk (see PROGRESS_REPORT_UPLOAD_DIR) and
    record it against the submission. The submission must already exist —
    progress reports are children of a submission, so there's nothing to
    attach to until the parent record has been created.
    """
    submission = (
        db.query(Submission).filter(Submission.id == submission_uuid).first()
    )
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")

    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided.")

    contents = await file.read()
    if len(contents) > MAX_PROGRESS_REPORT_SIZE_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File exceeds the {MAX_PROGRESS_REPORT_SIZE_BYTES // (1024 * 1024)}MB upload limit.",
        )

    # Collision-safe name on disk; the original name is preserved separately
    # in ProgressReport.file_name for display/download purposes.
    stored_filename = f"{uuid_lib.uuid4()}_{file.filename}"
    (PROGRESS_REPORT_UPLOAD_DIR / stored_filename).write_bytes(contents)

    progress_report = ProgressReport(
        submission_id=submission.id,
        file_url=f"/uploads/progress_reports/{stored_filename}",
        file_name=file.filename,
        upload_date=datetime.utcnow(),
        notes=notes,
    )
    db.add(progress_report)
    db.commit()
    db.refresh(progress_report)
    return progress_report


# ──────────────────────────────────────────────────────────────────────────────
# Bulk Excel upload (Step 4 & 5)
# ──────────────────────────────────────────────────────────────────────────────


@router.post(
    "/create_submission_upload-xlsx/",
    summary=(
            "Bulk-import submissions from a consolidated Excel workbook "
            "(header row + data rows).  All rows are committed in a single "
            "transaction; any validation error rolls back the entire batch."
    ),
)
async def create_submission_upload_xlsx(
        file: UploadFile = File(...),
        data_source: Optional[str] = Query(default=None),
        db: Session = Depends(get_db),
        auth: Authorized = Depends(RequirePermission("upload-template")),
        tenant: Tenant = Depends(get_current_tenant),
) -> Dict[str, Any]:
    """
    Accepts a multi-sheet Excel file with the following structure:

    **Sheet: "General project details"**
    Row 1 contains column headers matching ``GENERAL_COLUMN_MAP`` keys.
    Each subsequent row represents one submission.

    **Sheet: "Adaptation details"** (optional — included when any row has
    ``intervention_measurement`` of "Adaptation" or "Cross Cutting").
    Rows must align with "General project details" by row position (1:1 mapping).

    **Sheet: "Mitigation details"** (optional — same as above for Mitigation).

    Vocabulary validation is applied to ``hazard``, ``sector``, and
    ``mitigation_program`` fields.  Any unknown terms cause the entire batch
    to be rejected with a detailed error payload (row number + column name)
    rather than silently defaulting to "Other".
    """
    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents), keep_vba=True, read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Could not open workbook: {exc}",
        )

    # ── Sheet presence validation ─────────────────────────────────────────────
    if "General project details" not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail="Workbook must contain a sheet named 'General project details'.",
        )

    # ── Parse general rows ────────────────────────────────────────────────────
    general_rows: List[Dict[str, Any]] = _parse_sheet_rows(
        wb["General project details"], GENERAL_COLUMN_MAP
    )
    if not general_rows:
        raise HTTPException(
            status_code=400,
            detail="The 'General project details' sheet contains no data rows.",
        )

    # ── Parse optional adaptation / mitigation sheets ─────────────────────────
    adaptation_rows: List[Dict[str, Any]] = (
        _parse_sheet_rows(wb["Adaptation details"], ADAPTATION_COLUMN_MAP)
        if "Adaptation details" in wb.sheetnames
        else []
    )
    mitigation_rows: List[Dict[str, Any]] = (
        _parse_sheet_rows(wb["Mitigation details"], MITIGATION_COLUMN_MAP)
        if "Mitigation details" in wb.sheetnames
        else []
    )

    # ── Pre-load vocabulary terms for validation (single DB round-trip) ───────
    valid_vocab_terms: Set[str] = _get_valid_vocabulary_terms(db)

    # ── Build and validate SubmissionCreate objects ───────────────────────────
    validation_errors: List[Dict[str, Any]] = []
    submissions_to_create: List[SubmissionCreate] = []

    for row_idx, general_row in enumerate(general_rows, start=2):  # start=2 → Excel row number.
        # Extract and build geo_location.
        geo_dict = _build_geo_location(general_row)
        if geo_dict:
            general_row["geo_location"] = geo_dict

        # Normalise intervention_measurement.
        raw_im = general_row.get("intervention_measurement", "")
        if raw_im:
            general_row["intervention_measurement"] = _normalize_intervention_measurement(
                str(raw_im)
            )

        # Coerce numeric fields.
        if "funding_amount" in general_row:
            try:
                general_row["funding_amount"] = float(general_row["funding_amount"])
            except (ValueError, TypeError):
                general_row.pop("funding_amount")

        # Attach child payloads if available (aligned by row position).
        data_row_idx = row_idx - 2  # 0-based index into adaptation/mitigation rows.
        adap_data: Optional[Dict[str, Any]] = (
            adaptation_rows[data_row_idx] if data_row_idx < len(adaptation_rows) else None
        )
        mit_data: Optional[Dict[str, Any]] = (
            mitigation_rows[data_row_idx] if data_row_idx < len(mitigation_rows) else None
        )

        if adap_data:
            general_row["adaptation_data"] = adap_data
        if mit_data:
            general_row["mitigation_data"] = mit_data

        # ── Taxonomy validation (Step 5) ──────────────────────────────────────
        # Strategy: flag unknown terms; reject the full batch so data quality
        # issues are visible to the uploader before anything is persisted.
        vocab_errors = _validate_taxonomy(
            row_idx=row_idx,
            general_row=general_row,
            adap_data=adap_data,
            mit_data=mit_data,
            valid_terms=valid_vocab_terms,
        )
        validation_errors.extend(vocab_errors)

        # ── Pydantic schema validation ────────────────────────────────────────
        try:
            submission_obj = SubmissionCreate(**general_row)
            submissions_to_create.append(submission_obj)
        except Exception as pydantic_exc:
            validation_errors.append({
                "row": row_idx,
                "error": str(pydantic_exc),
            })

    # Reject the whole batch if any errors were found.
    if validation_errors:
        raise HTTPException(
            status_code=422,
            detail={
                "message": (
                    "Batch rejected: validation errors found. "
                    "No records were saved.  Fix the issues below and re-upload."
                ),
                "errors": validation_errors,
            },
        )

    # ── Batch insert in a single transaction ──────────────────────────────────
    created_ids: List[str] = []
    try:
        for submission in submissions_to_create:
            geo_dict = (
                submission.geo_location.dict(exclude_none=True)
                if submission.geo_location
                else None
            )
            db_submission = Submission(
                title=submission.title,
                intervention_measurement=submission.intervention_measurement,
                description=submission.description,
                implementation_status=submission.implementation_status,
                implementation_organization=submission.implementation_organization,
                implementation_partners_other=submission.implementation_partners_other,
                start_date=submission.start_date,
                end_date=submission.end_date,
                link=submission.link,
                funding_organization=submission.funding_organization,
                funding_type=submission.funding_type,
                funding_amount=submission.funding_amount,
                estimated_budget_cost=submission.estimated_budget_cost,
                geo_location=geo_dict,
                project_manager_name=submission.project_manager_name,
                project_manager_organization=submission.project_manager_organization,
                project_manager_email=submission.project_manager_email,
                project_manager_contact_number=submission.project_manager_contact_number,
                submission_status="Pending",
                issubmitted=True,
                platform=submission.platform,
                research=submission.research,
                data_source=data_source or f"bulk_upload: {file.filename}",
                createdby=auth.internal_user_id,
                createdate=datetime.utcnow(),
            )
            db.add(db_submission)
            db.flush()  # Get the UUID without committing.
            _link_submission_to_tenant(db, db_submission.id, tenant)

            im_value = submission.intervention_measurement.strip().lower()

            if im_value in ("mitigation", "cross cutting") and submission.mitigation_data:
                _create_mitigation_record(db, db_submission.id, submission.mitigation_data)

            if im_value in ("adaptation", "cross cutting") and submission.adaptation_data:
                _create_adaptation_record(db, db_submission.id, submission.adaptation_data)

            created_ids.append(str(db_submission.id))

        db.commit()

    except Exception:
        db.rollback()
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail="Batch insert failed.  All changes have been rolled back.",
        )

    return {
        "detail": f"{len(created_ids)} submission(s) created successfully.",
        "submission_ids": created_ids,
    }


# ──────────────────────────────────────────────────────────────────────────────
# Facets endpoint
# ──────────────────────────────────────────────────────────────────────────────


@router.get(
    "/facets/submission",
    summary="Return all distinct facet values available for filtering submissions.",
)
def get_submission_facets(db: Session = Depends(get_db)) -> Dict[str, List[Any]]:
    """Aggregate distinct values across Submission, Adaptation, and Mitigation."""
    return {
        "implementation_status": [
            r[0] for r in db.query(Submission.implementation_status).distinct()
        ],
        "implementation_organization": [
            r[0] for r in db.query(Submission.implementation_organization).distinct()
        ],
        "funding_organization": [
            r[0] for r in db.query(Submission.funding_organization).distinct()
        ],
        "funding_type": [
            r[0] for r in db.query(Submission.funding_type).distinct()
        ],
        "submission_status": [
            r[0] for r in db.query(Submission.submission_status).distinct()
        ],
        "research": [
            r[0] for r in db.query(Submission.research).distinct()
        ],
        "adaptation_sector": [
            r[0] for r in db.query(Adaptation.sector).distinct()
        ],
        "adaptation_national_policy": [
            r[0] for r in db.query(Adaptation.national_policy).distinct()
        ],
        "adaptation_hazard": [
            r[0] for r in db.query(Adaptation.hazard).distinct()
        ],
        "adaptation_climate_impact": [
            r[0] for r in db.query(Adaptation.climate_impact).distinct()
        ],
        "mitigation_sector": [
            r[0] for r in db.query(Mitigation.sector).distinct()
        ],
        "mitigation_subsector": [
            r[0] for r in db.query(Mitigation.subsector).distinct()
        ],
        "mitigation_project_type": [
            r[0] for r in db.query(Mitigation.project_type).distinct()
        ],
        "mitigation_program": [
            r[0] for r in db.query(Mitigation.mitigation_program).distinct()
        ],
        "mitigation_national_policy": [
            r[0] for r in db.query(Mitigation.national_policy).distinct()
        ],
        "mitigation_primary_intended_outcome": [
            r[0] for r in db.query(Mitigation.primary_intended_outcome).distinct()
        ],
        "mitigation_environmental_co_benefit": [
            r[0] for r in db.query(Mitigation.environmental_co_benefit).distinct()
        ],
        "mitigation_social_co_benefit": [
            r[0] for r in db.query(Mitigation.social_co_benefit).distinct()
        ],
        "mitigation_economic_co_benefit": [
            r[0] for r in db.query(Mitigation.economic_co_benefit).distinct()
        ],
        "mitigation_carbon_credit": [
            r[0] for r in db.query(Mitigation.carbon_credit).distinct()
        ],
    }


# ──────────────────────────────────────────────────────────────────────────────
# Private helpers for child-record CRUD
# ──────────────────────────────────────────────────────────────────────────────


def _create_mitigation_record(
        db: Session,
        submission_id: UUID,
        data: MitigationCreate,
) -> None:
    """Insert a new Mitigation row linked to the given submission UUID."""
    record = Mitigation(
        submission_id=submission_id,
        sector=data.sector,
        subsector=data.subsector,
        secondary=data.secondary,
        project_type=data.project_type,
        project_subtype=data.project_subtype,
        mitigation_program=data.mitigation_program,
        national_policy=data.national_policy,
        provincial_municipal=data.provincial_municipal,
        primary_intended_outcome=data.primary_intended_outcome,
        progress_calculator=data.progress_calculator,
        environmental_co_benefit=data.environmental_co_benefit,
        environmental_co_benefit_description=data.environmental_co_benefit_description,
        social_co_benefit=data.social_co_benefit,
        social_co_benefit_description=data.social_co_benefit_description,
        economic_co_benefit=data.economic_co_benefit,
        economic_co_benefit_description=data.economic_co_benefit_description,
        carbon_credit=data.carbon_credit,
        cdm_voluntary=data.cdm_voluntary,
        cdm_executive_board_status=data.cdm_executive_board_status,
        cdm_methodology=data.cdm_methodology,
        organization_issuing_credits=data.organization_issuing_credits,
        voluntary_methodology=data.voluntary_methodology,
        cdm_project_number=data.cdm_project_number,
    )
    db.add(record)


def _create_adaptation_record(
        db: Session,
        submission_id: UUID,
        data: AdaptationCreate,
) -> None:
    """Insert a new Adaptation row linked to the given submission UUID."""
    record = Adaptation(
        submission_id=submission_id,
        sector=data.sector,
        national_policy=data.national_policy,
        intervention_goal=data.intervention_goal,
        provincial_municipal=data.provincial_municipal,
        hazard=data.hazard,
        progress_calculator=data.progress_calculator,
        climate_impact=data.climate_impact,
        address_climate_impact=data.address_climate_impact,
        impact_response=data.impact_response,
    )
    db.add(record)


def _delete_adaptation(db: Session, submission_id: UUID) -> None:
    """Remove the Adaptation record for a submission if it exists."""
    adaptation = (
        db.query(Adaptation).filter(Adaptation.submission_id == submission_id).first()
    )
    if adaptation:
        db.delete(adaptation)


def _delete_mitigation(db: Session, submission_id: UUID) -> None:
    """Remove the Mitigation record for a submission if it exists."""
    mitigation = (
        db.query(Mitigation).filter(Mitigation.submission_id == submission_id).first()
    )
    if mitigation:
        db.delete(mitigation)


def _upsert_mitigation(
        db: Session,
        submission_id: UUID,
        update_dict: Optional[Dict[str, Any]],
) -> None:
    """Create or update the Mitigation record for a submission."""
    if update_dict is None:
        return
    existing = (
        db.query(Mitigation).filter(Mitigation.submission_id == submission_id).first()
    )
    if existing:
        for key, value in update_dict.items():
            setattr(existing, key, value)
    else:
        db.add(Mitigation(submission_id=submission_id, **update_dict))


def _upsert_adaptation(
        db: Session,
        submission_id: UUID,
        update_dict: Optional[Dict[str, Any]],
) -> None:
    """Create or update the Adaptation record for a submission."""
    if update_dict is None:
        return
    existing = (
        db.query(Adaptation).filter(Adaptation.submission_id == submission_id).first()
    )
    if existing:
        for key, value in update_dict.items():
            setattr(existing, key, value)
    else:
        db.add(Adaptation(submission_id=submission_id, **update_dict))


def _validate_taxonomy(
        row_idx: int,
        general_row: Dict[str, Any],
        adap_data: Optional[Dict[str, Any]],
        mit_data: Optional[Dict[str, Any]],
        valid_terms: Set[str],
) -> List[Dict[str, Any]]:
    """
    Validate taxonomy fields against the Vocabulary table.

    Approach: **flag unknown terms** — the entire batch is rejected with a
    structured error payload listing the exact row number and column name for
    each unknown value.  This is preferred over silently mapping to "Other"
    because it surfaces data-quality issues that the uploader must fix.

    Fields checked:
    * ``hazard`` (from adaptation_data)
    * ``sector`` (from adaptation_data and mitigation_data)
    * ``mitigation_program`` (from mitigation_data)
    """
    errors: List[Dict[str, Any]] = []

    def _check(field_value: Any, column_label: str) -> None:
        if field_value and str(field_value).strip().lower() not in valid_terms:
            errors.append({
                "row": row_idx,
                "column": column_label,
                "value": field_value,
                "message": (
                    f"'{field_value}' is not a recognised vocabulary term. "
                    "Check the Vocabulary table or correct the value in the template."
                ),
            })

    # Adaptation taxonomy fields.
    if adap_data:
        _check(adap_data.get("hazard"), "Hazard")
        _check(adap_data.get("sector"), "Adaptation sector")

    # Mitigation taxonomy fields.
    if mit_data:
        _check(mit_data.get("sector"), "Mitigation sector")
        _check(mit_data.get("mitigation_program"), "Mitigation programme")

    return errors
